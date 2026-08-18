# =====================================================================
#  SISTEM MARKING ANTI-DOUBLE - CV JAVA VOLUME ART (FINAL)
#  Konfirmasi hapus via pop-up modal (st.dialog)
#  Akun awal: admin / admin123
# =====================================================================
import io
import os
import re
import base64
import hashlib
import streamlit as st
import sqlite3
import pandas as pd
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

DB_FILE = "packing_java_volume_art.db"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

WARNA_OPSI = ["Natural Waterbase", "Black Burnt", "Black", "Dirty Brown",
              "Rustic", "White Bleached", "Bleached"]
LAINNYA = "Lainnya (ketik manual)"

SKALA_CETAK = 120   # persen

# ---------------- DETEKSI LOGO ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = None
for _nama in ("logo.png", "logo.jpg", "logo.jpeg", "logo.svg"):
    _c = os.path.join(BASE_DIR, _nama)
    if os.path.exists(_c):
        LOGO_PATH = _c
        break

LOGO_B64 = None
if LOGO_PATH:
    _ext = os.path.splitext(LOGO_PATH)[1][1:].lower()
    if _ext == "svg":
        _ext = "svg+xml"
    with open(LOGO_PATH, "rb") as _f:
        LOGO_B64 = f"data:image/{_ext};base64," + base64.b64encode(_f.read()).decode()

st.set_page_config(page_title="Marking CV Java Volume Art",
                   page_icon=LOGO_PATH if LOGO_PATH else "📦",
                   layout="wide", initial_sidebar_state="expanded")

# ================= TEMA =================
CUSTOM_CSS = """
<style>
:root{
  --coklat-tua:#6B4F3A; --coklat:#8A5A33; --emas:#C89F68;
  --krem:#FAF7F2; --krem-2:#F3EBDD; --garis:#E7DCCB; --teks:#4E342E;
}
html, body, [class*="css"]{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;}
#MainMenu, footer{visibility:hidden;}
.stApp{background:var(--krem);}

.brand{display:flex;align-items:center;gap:16px;padding:14px 20px;border-radius:18px;
  background:linear-gradient(135deg,var(--coklat-tua) 0%,#8A6A50 55%,var(--emas) 100%);
  color:#fff;box-shadow:0 6px 18px rgba(93,64,55,.25);margin-bottom:14px;}
.brand .logo{font-size:30px;background:rgba(255,255,255,.16);border-radius:14px;padding:6px 10px;}
.brand .logo-img{height:58px;width:auto;background:#fff;border-radius:12px;padding:4px 10px;object-fit:contain;}
.brand h1{margin:0;font-size:24px;font-weight:800;color:#fff !important;letter-spacing:.3px;}
.brand p{margin:3px 0 0;font-size:12px;color:#F3E7D8;}

h1,h2,h3,h4{color:var(--teks) !important;font-weight:800 !important;}
p,span,label{color:var(--teks);}
hr{border-color:var(--garis) !important;}

[data-testid="stVerticalBlockBorder"]{border-radius:18px !important;border:1px solid var(--garis) !important;
  background:#fff;box-shadow:0 3px 12px rgba(93,64,55,.08);}

button[data-testid="stBaseButton-primary"], .stButton>button[kind="primary"]{
  background:var(--coklat) !important;color:#fff !important;border:none !important;
  border-radius:12px !important;font-weight:700 !important;padding:10px 24px !important;
  box-shadow:0 4px 12px rgba(138,90,51,.35) !important;}
button[data-testid="stBaseButton-primary"]:hover{background:var(--coklat-tua) !important;}
button[data-testid="stBaseButton-secondary"], .stDownloadButton>button{
  border-radius:12px !important;border:1.5px solid var(--emas) !important;
  color:var(--coklat-tua) !important;font-weight:600 !important;background:#fff !important;}

input, textarea{border-radius:10px !important;border:1.5px solid #DCCDB6 !important;
  background:#FFFDF9 !important;}
input:focus, textarea:focus{border-color:var(--coklat) !important;
  box-shadow:0 0 0 3px rgba(200,159,104,.25) !important;outline:none !important;}
input[type="checkbox"], input[type="radio"]{accent-color:var(--coklat);width:16px;height:16px;}

section[data-testid="stSidebar"]{background:var(--krem-2);}
[data-testid="stMetric"]{background:#fff;border:1px solid var(--garis);border-radius:14px;padding:10px 14px;}
.stAlert{border-radius:12px !important;}
[data-testid="stDataFrame"]{border-radius:12px;overflow:hidden;}

section[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"]{
  opacity:0;width:0;height:0;margin:0;position:absolute;}
section[data-testid="stSidebar"] [data-testid="stRadio"] label{
  display:block;padding:10px 14px;border-radius:12px;font-weight:600;cursor:pointer;margin:2px 0;}
section[data-testid="stSidebar"] [data-testid="stRadio"] label:hover{background:#EADFCB;}
section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked){
  background:var(--coklat);box-shadow:0 3px 8px rgba(138,90,51,.3);}
section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) span,
section[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) p{
  color:#fff !important;}
</style>
"""

if LOGO_B64:
    LOGO_HTML = f'<img src="{LOGO_B64}" class="logo-img" alt="Logo JVA">'
else:
    LOGO_HTML = '<div class="logo">📦</div>'

BRAND_HTML = f"""
<div class="brand">
  {LOGO_HTML}
  <div>
    <h1>CV Java Volume Art</h1>
    <p>Sistem Marking Anti-Double &bull; Furniture &amp; Home Decoration</p>
  </div>
</div>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ---------------- KEAMANAN KATA SANDI ----------------
def hash_password(password, salt=None):
    salt = salt or os.urandom(16)
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100_000)
    return salt.hex(), h.hex()

def verifikasi_password(password, salt_hex, hash_hex):
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"),
                            bytes.fromhex(salt_hex), 100_000)
    return h.hex() == hash_hex

# ---------------- DATABASE ----------------
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS packing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            packing_no TEXT NOT NULL,
            shipment_no TEXT NOT NULL,
            code TEXT, description TEXT, size TEXT,
            colour TEXT, qty TEXT,
            replacement INTEGER NOT NULL DEFAULT 0,
            created_at TEXT, created_by TEXT,
            UNIQUE (shipment_no, packing_no)
        )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS warna_custom (nama TEXT PRIMARY KEY)""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'staff',
            created_at TEXT
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reprint_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shipment_no TEXT, packing_no TEXT, alasan TEXT,
            dicetak_oleh TEXT, created_at TEXT
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aksi TEXT, shipment_no TEXT, packing_no TEXT,
            detail TEXT, dilakukan_oleh TEXT, created_at TEXT
        )""")
    cols = [c[1] for c in conn.execute("PRAGMA table_info(packing)").fetchall()]
    if "replacement" not in cols:
        conn.execute("ALTER TABLE packing ADD COLUMN replacement INTEGER NOT NULL DEFAULT 0")
    if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        s, h = hash_password("admin123")
        conn.execute("INSERT INTO users (username,password_hash,salt,role,created_at) "
                     "VALUES (?,?,?,?,?)",
                     ("admin", h, s, "admin", datetime.now().strftime("%d-%m-%Y %H:%M")))
    conn.commit()
    return conn

def log_audit(conn, aksi, shipment_no, packing_no, detail=""):
    conn.execute("INSERT INTO audit_log "
                 "(aksi,shipment_no,packing_no,detail,dilakukan_oleh,created_at) "
                 "VALUES (?,?,?,?,?,?)",
                 (aksi, shipment_no, packing_no, detail,
                  st.session_state.get("login", ""),
                  datetime.now().strftime("%d-%m-%Y %H:%M")))

# ---------------- POP-UP KONFIRMASI HAPUS ----------------
@st.dialog("⚠️ Konfirmasi Hapus")
def konfirmasi_hapus_ship(ship, n_koli):
    st.markdown(f"Shipment **{ship}** berisi **{n_koli} koli** akan dihapus **permanen**.")
    st.caption("Tindakan ini tidak dapat dibatalkan dan tercatat di riwayat audit.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ Ya, Hapus Semua", type="primary"):
            conn = get_db()
            conn.execute("DELETE FROM packing WHERE shipment_no = ?", (ship,))
            log_audit(conn, "HAPUS SHIPMENT", ship, "-", f"{n_koli} koli dihapus")
            conn.commit(); conn.close()
            st.session_state.pop("konfirmasi_hapus", None)
            st.rerun()
    with c2:
        if st.button("Batal"):
            st.session_state.pop("konfirmasi_hapus", None)
            st.rerun()

@st.dialog("⚠️ Konfirmasi Hapus")
def konfirmasi_hapus_pack(row):
    st.markdown(f"Nomor **{row['packing_no']}** (shipment **{row['shipment_no']}**) akan "
                f"dihapus **permanen** dan nomornya tidak akan dipakai ulang.")
    st.caption("Tindakan ini tidak dapat dibatalkan dan tercatat di riwayat audit.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ Ya, Hapus", type="primary"):
            conn = get_db()
            conn.execute("DELETE FROM packing WHERE id = ?", (row["id"],))
            log_audit(conn, "HAPUS LABEL", row["shipment_no"], row["packing_no"],
                      f"description: {row['description']}")
            conn.commit(); conn.close()
            st.session_state.pop("konfirmasi_hapus", None)
            st.rerun()
    with c2:
        if st.button("Batal"):
            st.session_state.pop("konfirmasi_hapus", None)
            st.rerun()

def daftar_warna():
    conn = get_db()
    extra = [r["nama"] for r in conn.execute(
             "SELECT nama FROM warna_custom ORDER BY nama").fetchall()]
    conn.close()
    return WARNA_OPSI + extra

def daftar_kode_buyer():
    conn = get_db()
    ships = [r["shipment_no"] for r in conn.execute(
             "SELECT DISTINCT shipment_no FROM packing").fetchall()]
    conn.close()
    kode = set()
    for s in ships:
        m = re.match(r"[A-Za-z]+", s or "")
        kode.add(m.group(0) if m else (s or "").strip())
    return sorted(k for k in kode if k)

def kunci_urut(no):
    return [int(x) for x in re.split(r"[.\-]", str(no)) if x.isdigit()]

def parse_packing(no):
    bagian = [b for b in re.split(r"[.\-]", str(no).strip()) if b.strip().isdigit()]
    return pd.Series({"utama": int(bagian[0]) if len(bagian) > 0 else 0,
                      "sub":   int(bagian[1]) if len(bagian) > 1 else 0})

def next_range(conn, shipment_no, nomor_utama, jumlah=1):
    cur = conn.execute(
        "SELECT packing_no FROM packing WHERE shipment_no = ? AND packing_no LIKE ?",
        (shipment_no, f"{nomor_utama}.%"))
    subs = []
    for row in cur.fetchall():
        try: subs.append(int(row["packing_no"].split(".")[1]))
        except (IndexError, ValueError): continue
    start = (max(subs) if subs else 0) + 1
    return [f"{nomor_utama}.{start + i}" for i in range(jumlah)]

def simpan_batch(conn, items):
    try:
        conn.executemany("""INSERT INTO packing
            (packing_no,shipment_no,code,description,size,colour,qty,replacement,created_at)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            [(d["packing_no"], d["shipment_no"], d["code"], d["description"],
              d["size"], d["colour"], d["qty"], d["replacement"],
              datetime.now().strftime("%d-%m-%Y %H:%M")) for d in items])
        conn.commit()
        return True, None
    except sqlite3.IntegrityError:
        conn.rollback()
        return False, "Sebagian nomor ternyata sudah diterbitkan pengguna lain."

# ---------------- EXPORT EXCEL ----------------
def buat_excel_template(items):
    wb = Workbook(); ws = wb.active; ws.title = "Packing List"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 45
    TEBAL = Side(style="medium", color="000000")
    TIPIS = Side(style="thin",   color="B0B0B0")

    prev_ship, count_in_ship = None, 0
    for idx, d in enumerate(items):
        r = 1 + idx * 10
        if d["shipment_no"] != prev_ship:
            if idx > 0:
                ws.row_breaks.append(Break(id=r - 1))
            prev_ship, count_in_ship = d["shipment_no"], 0
        elif count_in_ship % 2 == 0:
            ws.row_breaks.append(Break(id=r - 1))
        count_in_ship += 1

        for i in range(9):
            ws.row_dimensions[r + i].height = 22 if i < 3 else 24
        ws.row_dimensions[r + 9].height = 8

        for row in range(r, r + 9):
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = Border(
                    left  = TEBAL if col == 1 else TIPIS,
                    right = TEBAL if col == 3 else TIPIS,
                    top   = TEBAL if row == r else TIPIS,
                    bottom= TEBAL if row == r + 8 else TIPIS)

        ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r+2, end_column=3)

        ws[f"A{r}"] = "PACKING NO"
        ws[f"A{r}"].font = Font(name="Arial", size=13)
        ws[f"A{r}"].alignment = Alignment(vertical="center")
        ws[f"B{r}"] = ":"
        ws[f"C{r}"] = d["packing_no"]
        ws[f"C{r}"].font = Font(name="Arial", size=36, bold=True)
        ws[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")

        for i, (k, v) in enumerate([("SHIPMENT NO", d["shipment_no"]), ("CODE", d["code"]),
                                    ("DESCRIPTION", d["description"]), ("SIZE", d["size"]),
                                    ("COLOUR", d["colour"]), ("QTY", d["qty"])]):
            ws[f"A{r+3+i}"] = k
            ws[f"A{r+3+i}"].font = Font(name="Arial", size=13)
            ws[f"A{r+3+i}"].alignment = Alignment(vertical="center")
            ws[f"B{r+3+i}"] = ":"
            ws[f"C{r+3+i}"] = v
            ws[f"C{r+3+i}"].font = Font(name="Arial", size=13, bold=True)
            ws[f"C{r+3+i}"].alignment = Alignment(vertical="center")

        if d.get("replacement"):
            ws[f"C{r+3}"] = CellRichText(
                TextBlock(InlineFont(rFont="Arial", sz=13, b=True), str(d["shipment_no"])),
                TextBlock(InlineFont(rFont="Arial", sz=13, b=True, color="FF0000"),
                          "  # REPLACEMENT"))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ---------------- LABEL HTML ----------------
def html_label(d):
    ship = str(d["shipment_no"])
    if d.get("replacement"):
        ship += ' <span style="color:#d40000;font-weight:bold"># REPLACEMENT</span>'
    rows = "".join(
        f'<tr><td style="border:2px solid #000;padding:6px 10px;width:140px;font-family:Arial">{k}</td>'
        f'<td style="border:2px solid #000;padding:6px 10px;font-family:Arial;font-weight:bold">{v}</td></tr>'
        for k, v in [("SHIPMENT NO", ship), ("CODE", d["code"]),
                     ("DESCRIPTION", d["description"]), ("SIZE", d["size"]),
                     ("COLOUR", d["colour"]), ("QTY", d["qty"])])
    return ('<table style="border-collapse:collapse;width:430px">'
            '<tr><td style="border:2px solid #000;padding:6px 10px;width:140px;font-family:Arial">PACKING NO</td>'
            f'<td style="border:2px solid #000;padding:6px;font-family:Arial;font-size:46px;font-weight:bold;'
            f'text-align:center">{d["packing_no"]}</td></tr>{rows}</table>')

def html_batch(items, auto_print=False):
    blok = "".join(f'<div style="margin-bottom:14px;page-break-inside:avoid">{html_label(d)}</div>'
                   for d in items)
    if auto_print:
        return (f"<html><head><style>body{{zoom:{SKALA_CETAK / 100};}}</style></head>"
                f"<body onload='window.print()'>{blok}</body></html>")
    return blok

# ================= BARIS ATAS =================
hb_kiri, hb_kanan = st.columns([7, 1])
with hb_kiri:
    st.markdown(BRAND_HTML, unsafe_allow_html=True)
with hb_kanan:
    if st.session_state.get("login"):
        with st.popover(f"👤 {st.session_state['login']}"):
            st.caption(f"Peran: **{st.session_state.get('role')}**")

            st.markdown("**🔑 Ganti Kata Sandi**")
            pl = st.text_input("Kata sandi lama", type="password")
            pb = st.text_input("Kata sandi baru (min. 6 karakter)", type="password")
            if st.button("Ganti Kata Sandi", key="btn_ganti"):
                conn = get_db()
                me = conn.execute("SELECT * FROM users WHERE username = ?",
                                  (st.session_state["login"],)).fetchone()
                if not verifikasi_password(pl, me["salt"], me["password_hash"]):
                    st.error("Kata sandi lama salah.")
                elif len(pb) < 6:
                    st.error("Kata sandi baru minimal 6 karakter.")
                else:
                    s, h = hash_password(pb)
                    conn.execute("UPDATE users SET password_hash = ?, salt = ? WHERE username = ?",
                                 (h, s, st.session_state["login"]))
                    conn.commit()
                    st.success("Kata sandi berhasil diganti.")
                conn.close()

            if st.session_state.get("role") == "admin":
                st.divider()
                st.markdown("**👥 Manajemen Pengguna**")
                nu = st.text_input("Username baru")
                npw = st.text_input("Kata sandi", type="password")
                role_baru = st.selectbox("Peran", ["staff", "admin"])
                if st.button("Tambah Pengguna", key="btn_tambah"):
                    if not (nu.strip() and npw):
                        st.error("Username dan kata sandi wajib diisi.")
                    elif len(npw) < 6:
                        st.error("Kata sandi minimal 6 karakter.")
                    else:
                        s, h = hash_password(npw)
                        conn = get_db()
                        try:
                            conn.execute("INSERT INTO users "
                                         "(username,password_hash,salt,role,created_at) "
                                         "VALUES (?,?,?,?,?)",
                                         (nu.strip(), h, s, role_baru,
                                          datetime.now().strftime("%d-%m-%Y %H:%M")))
                            conn.commit()
                            st.success(f"Pengguna **{nu.strip()}** ditambahkan.")
                        except sqlite3.IntegrityError:
                            st.error("Username sudah dipakai.")
                        conn.close()
                conn = get_db()
                st.caption("Terdaftar: " + ", ".join(
                    r["username"] for r in conn.execute(
                        "SELECT username FROM users ORDER BY username").fetchall()))
                conn.close()

            st.divider()
            if st.button("🚪 Keluar", key="btn_keluar"):
                for k in ["login", "role", "batch", "reprint_item", "kelola", "konfirmasi_hapus"]:
                    st.session_state.pop(k, None)
                st.rerun()

# ================= HALAMAN LOGIN =================
if "login" not in st.session_state:
    st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("<h2 style='text-align:center'>🔐 Masuk Karyawan</h2>",
                    unsafe_allow_html=True)
        st.markdown("<p style='text-align:center'>Sistem Marking Anti-Double<br>"
                    "<b>CV Java Volume Art</b></p>", unsafe_allow_html=True)
        k1, k2, k3 = st.columns([1, 2, 1])
        with k2:
            u = st.text_input("Nama pengguna")
            p = st.text_input("Kata sandi", type="password")
            if st.button("🔓 Masuk", type="primary"):
                if not (u.strip() and p):
                    st.error("Nama pengguna dan kata sandi wajib diisi.")
                else:
                    conn = get_db()
                    row = conn.execute("SELECT * FROM users WHERE username = ?",
                                       (u.strip(),)).fetchone()
                    conn.close()
                    if row and verifikasi_password(p, row["salt"], row["password_hash"]):
                        st.session_state["login"] = row["username"]
                        st.session_state["role"] = row["role"]
                        st.rerun()
                    else:
                        st.error("Nama pengguna atau kata sandi salah.")
            conn = get_db()
            hanya_satu = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 1
            conn.close()
            if hanya_satu:
                st.caption("Akun awal: **admin / admin123** — segera ganti setelah masuk.")
    st.stop()

# ---------------- SIDEBAR ----------------
conn = get_db()
total, = conn.execute("SELECT COUNT(*) FROM packing").fetchone()
conn.close()
st.sidebar.markdown("### 📦 Navigasi")
st.sidebar.metric("Total label tersimpan", total)
menu = st.sidebar.radio("Menu", ["🏷️ Buat Label Baru", "📋 Daftar Packing",
                                 "🖨️ Cetak Ulang", "🔍 Audit Duplikat"],
                        label_visibility="collapsed")
st.sidebar.caption("© 2026 CV Java Volume Art — pemakaian internal")

# ---------------- MENU 1: BUAT LABEL ----------------
if menu == "🏷️ Buat Label Baru":
    st.markdown("#### Buat Packing No Baru")
    with st.container(border=True):
        m1, m2, m3 = st.columns([1, 1, 2], gap="small")
        with m1:
            mode = st.radio("Mode", ["Satuan (1 label)", "Borongan (banyak sekaligus)"], horizontal=True)
        with m2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            repl = st.checkbox("REPLACEMENT (penggantian barang)")
        if repl:
            st.markdown('Label akan ditandai: '
                        '<span style="color:#d40000;font-weight:bold"># REPLACEMENT</span>',
                        unsafe_allow_html=True)
        jumlah = 45 if mode.startswith("Borongan") else 1

        c1, c2 = st.columns(2)
        with c1:
            shipment = st.text_input("Shipment No", placeholder="EWS 020/VIII/2026")
            utama    = st.text_input("Nomor Utama", placeholder="10")
            if mode.startswith("Borongan"):
                jumlah = st.number_input("Jumlah koli yang akan dibuat", 1, 500, 45)
            if shipment.strip() and utama.strip() and utama.strip().isdigit():
                conn = get_db()
                rentang = next_range(conn, shipment.strip(), utama.strip(), jumlah)
                conn.close()
                st.success(f"Akan diterbitkan: **{rentang[0]} s/d {rentang[-1]}** ({jumlah} label)"
                           if jumlah > 1 else f"Nomor berikutnya: **{rentang[0]}**")
        with c2:
            code = st.text_input("Code", placeholder="#10")
            desc = st.text_input("Description", placeholder="DAISY BIKINI WOODEN DUCK")
            size = st.text_input("Size", placeholder="20x17x45 CM")

            pilih_warna = st.selectbox("Colour", daftar_warna() + [LAINNYA])
            if pilih_warna == LAINNYA:
                colour = st.text_input("Ketik warna baru", placeholder="cth: YELLOW STABILO")
                if colour.strip():
                    st.caption(f"Warna baru **“{colour.strip()}”** akan masuk daftar pilihan.")
            else:
                colour = pilih_warna

            qty = st.text_input("Qty", placeholder="4 PCS")

        if st.button(f"🔒 Terbitkan {jumlah} Label Sekaligus" if jumlah > 1
                     else "🔒 Terbitkan & Buat Label", type="primary"):
            if not (shipment.strip() and utama.strip()):
                st.error("Shipment No dan Nomor Utama wajib diisi.")
            elif not utama.strip().isdigit():
                st.error("Nomor Utama harus berupa angka, mis. 10 (tanpa strip/titik).")
            elif pilih_warna == LAINNYA and not colour.strip():
                st.error("Anda memilih 'Lainnya' — silakan ketik nama warnanya dahulu.")
            else:
                colour = colour.strip()
                conn = get_db()
                if pilih_warna == LAINNYA:
                    conn.execute("INSERT OR IGNORE INTO warna_custom (nama) VALUES (?)", (colour,))
                    conn.commit()
                rentang = next_range(conn, shipment.strip(), utama.strip(), jumlah)
                items = [{"packing_no": no, "shipment_no": shipment.strip(), "code": code,
                          "description": desc, "size": size, "colour": colour, "qty": qty,
                          "replacement": 1 if repl else 0}
                         for no in rentang]
                ok, err = simpan_batch(conn, items)
                conn.close()
                if ok:
                    st.session_state["batch"] = items
                    st.success(f"{len(items)} label berhasil diterbitkan: {rentang[0]} s/d {rentang[-1]}.")
                else:
                    st.error(f"{err} Muat ulang halaman dan coba lagi.")

    if "batch" in st.session_state:
        items = st.session_state["batch"]
        with st.container(border=True):
            st.markdown(f"#### 🖨️ {len(items)} Label Siap Cetak")
            st.dataframe(pd.DataFrame(items), hide_index=True)
            with st.expander("Pratinjau semua label"):
                st.markdown(html_batch(items), unsafe_allow_html=True)
            b1, b2 = st.columns(2)
            with b1:
                st.download_button("Unduh HTML Cetak Cepat",
                                   data=html_batch(items, auto_print=True),
                                   file_name=f"label_{items[0]['packing_no']}-{items[-1]['packing_no']}.html",
                                   mime="text/html")
            with b2:
                st.download_button("Export Excel (format template)",
                                   data=buat_excel_template(items).read(),
                                   file_name=f"packing_{items[0]['packing_no']}-{items[-1]['packing_no']}.xlsx",
                                   mime=MIME_XLSX)

# ---------------- MENU 2: DAFTAR + KELOLA DATA ----------------
elif menu == "📋 Daftar Packing":
    st.markdown("#### Daftar Packing No per Shipment")
    with st.container(border=True):
        st.caption("Pilih kode buyer — seluruh shipment dengan kode tersebut langsung tampil.")
        kode_tersedia = daftar_kode_buyer()
        MANUAL = "— ketik manual —"
        pilih_kode = st.selectbox("Kode Buyer", ([MANUAL] if not kode_tersedia
                                                 else kode_tersedia + [MANUAL]))
        if pilih_kode == MANUAL:
            cari = st.text_input("Awalan Shipment No", placeholder="cth: DGE")
            kondisi, params, aktif = "shipment_no LIKE ?", (cari.strip() + "%",), bool(cari.strip())
            nama_file = cari.strip() or "semua"
        else:
            cari = pilih_kode
            kondisi, params, aktif = ("(shipment_no LIKE ? || '/%' OR shipment_no LIKE ? || ' %')",
                                      (pilih_kode, pilih_kode), True)
            nama_file = pilih_kode
        filter_warna = st.selectbox("Filter warna", ["Semua warna"] + daftar_warna())

    if aktif:
        conn = get_db()
        rows = conn.execute(f"SELECT id,packing_no,shipment_no,code,description,size,colour,qty,replacement "
                            f"FROM packing WHERE {kondisi}", params).fetchall()
        conn.close()
        if rows:
            df = pd.DataFrame([dict(r) for r in rows])
            df[["utama", "sub"]] = df["packing_no"].apply(parse_packing)
            df = df.sort_values(["shipment_no", "utama", "sub"]).drop(columns=["utama", "sub"])
            if filter_warna != "Semua warna":
                df = df[df["colour"].fillna("").str.strip().str.lower() == filter_warna.lower()]
            if df.empty:
                st.info(f"Tidak ada label dengan warna **{filter_warna}** untuk kode **{cari}**.")
            else:
                df_tampil = df.drop(columns=["id"])
                with st.container(border=True):
                    col_ring, col_kelola = st.columns([3, 2], gap="large")
                    with col_ring:
                        st.markdown("#### Ringkasan per shipment")
                        st.dataframe(df_tampil.groupby("shipment_no").size()
                                       .reset_index(name="jumlah koli"), hide_index=True)
                    with col_kelola:
                        st.markdown("#### 🛠️ Kelola Data")
                        st.markdown("**🚚 Per Shipment**")
                        ship_ops = sorted(df_tampil["shipment_no"].unique().tolist())
                        ship_target = st.selectbox("Shipment", ship_ops, key="ship_target")
                        s1, s2 = st.columns(2)
                        with s1:
                            if st.button("✏️ Edit", key="btn_edit_ship"):
                                st.session_state["kelola"] = {"jenis": "edit_ship", "ship": ship_target}
                                st.session_state.pop("konfirmasi_hapus", None)
                        with s2:
                            if st.button("🗑️ Hapus", key="btn_hapus_ship"):
                                st.session_state["konfirmasi_hapus"] = {
                                    "tipe": "ship", "ship": ship_target,
                                    "n": int(len(df[df["shipment_no"] == ship_target]))}
                                st.session_state.pop("kelola", None)

                        st.markdown("**🏷️ Per Packing No**")
                        df_ship = df[df["shipment_no"] == ship_target]
                        pack_target = st.selectbox("Packing No", df_ship["packing_no"].tolist(),
                                                   key="pack_target")
                        row_sel = df_ship[df_ship["packing_no"] == pack_target].iloc[0]
                        p1, p2 = st.columns(2)
                        with p1:
                            if st.button("✏️ Edit", key="btn_edit_pack"):
                                st.session_state["kelola"] = {"jenis": "edit_pack", "row": dict(row_sel)}
                                st.session_state.pop("konfirmasi_hapus", None)
                        with p2:
                            if st.button("🗑️ Hapus", key="btn_hapus_pack"):
                                st.session_state["konfirmasi_hapus"] = {"tipe": "pack", "row": dict(row_sel)}
                                st.session_state.pop("kelola", None)

                    st.download_button("Export Hasil ke Excel",
                                       data=buat_excel_template(df_tampil.to_dict("records")).read(),
                                       file_name=f"packing_{nama_file}.xlsx",
                                       mime=MIME_XLSX)

                # ===== PANEL EDIT (inline) =====
                kelola = st.session_state.get("kelola")
                if kelola:
                    if kelola["jenis"] == "edit_ship":
                        with st.container(border=True):
                            st.markdown(f"#### ✏️ Edit Shipment {kelola['ship']}")
                            with st.form("form_edit_ship"):
                                baru = st.text_input("Shipment No baru", value=kelola["ship"])
                                if st.form_submit_button("💾 Simpan Perubahan", type="primary"):
                                    if not baru.strip():
                                        st.error("Shipment No baru wajib diisi.")
                                    else:
                                        conn = get_db()
                                        try:
                                            conn.execute("UPDATE packing SET shipment_no = ? "
                                                         "WHERE shipment_no = ?",
                                                         (baru.strip(), kelola["ship"]))
                                            log_audit(conn, "EDIT SHIPMENT", baru.strip(), "-",
                                                      f"dari: {kelola['ship']}")
                                            conn.commit(); conn.close()
                                            st.session_state.pop("kelola", None)
                                            st.success("Shipment diperbarui.")
                                            st.rerun()
                                        except sqlite3.IntegrityError:
                                            conn.rollback(); conn.close()
                                            st.error("Gagal: sebagian nomor sudah ada di "
                                                     "shipment tujuan.")
                            if st.button("Tutup", key="tutup_es"):
                                st.session_state.pop("kelola", None); st.rerun()

                    elif kelola["jenis"] == "edit_pack":
                        ei = kelola["row"]
                        with st.container(border=True):
                            st.markdown(f"#### ✏️ Edit Label {ei['packing_no']}")
                            with st.form(f"form_edit_{ei['id']}"):
                                f1, f2 = st.columns(2)
                                with f1:
                                    e_pack = st.text_input("Packing No", value=str(ei["packing_no"]))
                                    e_ship = st.text_input("Shipment No", value=str(ei["shipment_no"]))
                                    e_code = st.text_input("Code", value=str(ei["code"] or ""))
                                    e_desc = st.text_input("Description", value=str(ei["description"] or ""))
                                with f2:
                                    e_size = st.text_input("Size", value=str(ei["size"] or ""))
                                    e_col  = st.text_input("Colour", value=str(ei["colour"] or ""))
                                    e_qty  = st.text_input("Qty", value=str(ei["qty"] or ""))
                                    e_repl = st.checkbox("REPLACEMENT", value=bool(ei["replacement"]))
                                if st.form_submit_button("💾 Simpan Perubahan", type="primary"):
                                    if not (e_pack.strip() and e_ship.strip()):
                                        st.error("Packing No dan Shipment No wajib diisi.")
                                    else:
                                        conn = get_db()
                                        try:
                                            conn.execute("""UPDATE packing SET
                                                packing_no=?, shipment_no=?, code=?, description=?,
                                                size=?, colour=?, qty=?, replacement=?
                                                WHERE id=?""",
                                                (e_pack.strip(), e_ship.strip(), e_code.strip(),
                                                 e_desc.strip(), e_size.strip(), e_col.strip(),
                                                 e_qty.strip(), 1 if e_repl else 0, ei["id"]))
                                            log_audit(conn, "EDIT LABEL", e_ship.strip(), e_pack.strip(),
                                                      f"data lama: {ei['packing_no']} | {ei['description']}")
                                            conn.commit(); conn.close()
                                            st.session_state.pop("kelola", None)
                                            st.success("Perubahan disimpan.")
                                            st.rerun()
                                        except sqlite3.IntegrityError:
                                            conn.rollback(); conn.close()
                                            st.error("Gagal: kombinasi Shipment + Packing No "
                                                     "tersebut sudah dipakai label lain.")
                            if st.button("Tutup", key="tutup_ep"):
                                st.session_state.pop("kelola", None); st.rerun()

                # ===== RIWAYAT =====
                conn = get_db()
                alogs = conn.execute("SELECT created_at, aksi, shipment_no, packing_no, "
                                     "dilakukan_oleh FROM audit_log ORDER BY id DESC "
                                     "LIMIT 10").fetchall()
                conn.close()
                if alogs:
                    with st.expander("📒 Riwayat Edit / Hapus (10 terakhir)"):
                        st.dataframe(pd.DataFrame([dict(r) for r in alogs]), hide_index=True)
        else:
            st.info(f"Belum ada shipment yang berawalan **{cari}**.")

# ---------------- MENU 3: CETAK ULANG ----------------
elif menu == "🖨️ Cetak Ulang":
    st.markdown("#### Cetak Ulang Label (Sobek / Rusak / Hilang)")
    with st.container(border=True):
        st.caption("Cetak ulang **tidak membuat nomor ganda** — data diambil dari label yang "
                   "sudah tersimpan, dan kejadiannya dicatat di log sebagai bahan evaluasi.")
        kode_tersedia = daftar_kode_buyer()
        MANUAL = "— ketik manual —"
        pk = st.selectbox("Kode Buyer", ([MANUAL] if not kode_tersedia
                                         else kode_tersedia + [MANUAL]))
        if pk == MANUAL:
            cari_ship = st.text_input("Awalan Shipment No", placeholder="cth: EWS")
            conn = get_db()
            ships = [r["shipment_no"] for r in conn.execute(
                "SELECT DISTINCT shipment_no FROM packing WHERE shipment_no LIKE ? "
                "ORDER BY shipment_no", (cari_ship.strip() + "%",)).fetchall()]
            conn.close()
        else:
            conn = get_db()
            ships = [r["shipment_no"] for r in conn.execute(
                "SELECT DISTINCT shipment_no FROM packing WHERE "
                "(shipment_no LIKE ? || '/%' OR shipment_no LIKE ? || ' %') "
                "ORDER BY shipment_no", (pk, pk)).fetchall()]
            conn.close()

        if not ships:
            st.info("Tidak ada shipment yang cocok.")
        else:
            ship_pilih = st.selectbox("Shipment", ships)
            conn = get_db()
            rows = conn.execute("SELECT * FROM packing WHERE shipment_no = ?",
                                (ship_pilih,)).fetchall()
            conn.close()
            rows = sorted(rows, key=lambda r: kunci_urut(r["packing_no"]))
            nums = [r["packing_no"] for r in rows]
            pilih_no = st.selectbox("Packing No yang sobek / akan dicetak ulang", nums)
            alasan = st.text_input("Alasan cetak ulang", value="Label sobek")
            if st.button(f"✅ Yakin cetak ulang {pilih_no}?", type="primary"):
                row = rows[nums.index(pilih_no)]
                conn = get_db()
                conn.execute("INSERT INTO reprint_log "
                             "(shipment_no,packing_no,alasan,dicetak_oleh,created_at) "
                             "VALUES (?,?,?,?,?)",
                             (ship_pilih, pilih_no, alasan,
                              st.session_state.get("login", ""),
                              datetime.now().strftime("%d-%m-%Y %H:%M")))
                conn.commit(); conn.close()
                st.session_state["reprint_item"] = dict(row)
                st.success(f"Nomor **{pilih_no}** disiapkan. Silakan unduh dan cetak.")

    if "reprint_item" in st.session_state:
        item = st.session_state["reprint_item"]
        with st.container(border=True):
            st.markdown(f"#### 🖨️ Label {item['packing_no']} Siap Dicetak")
            st.markdown(html_label(item), unsafe_allow_html=True)
            b1, b2 = st.columns(2)
            with b1:
                st.download_button("Unduh HTML Cetak",
                                   data=html_batch([item], auto_print=True),
                                   file_name=f"label_{item['packing_no']}.html",
                                   mime="text/html")
            with b2:
                st.download_button("Export Excel",
                                   data=buat_excel_template([item]).read(),
                                   file_name=f"label_{item['packing_no']}.xlsx",
                                   mime=MIME_XLSX)

    conn = get_db()
    logs = conn.execute("SELECT created_at, shipment_no, packing_no, alasan, dicetak_oleh "
                        "FROM reprint_log ORDER BY id DESC LIMIT 10").fetchall()
    conn.close()
    if logs:
        with st.container(border=True):
            st.markdown("#### 📒 Log Cetak Ulang (10 terakhir)")
            st.dataframe(pd.DataFrame([dict(r) for r in logs]), hide_index=True)

# ---------------- MENU 4: AUDIT ----------------
elif menu == "🔍 Audit Duplikat":
    st.markdown("#### Audit Duplikat Data Lama")
    with st.container(border=True):
        st.caption("Tempel satu kolom packing no dari Excel (satu nomor per baris).")
        teks = st.text_area("Data", height=250, placeholder="10.1\n10.2\n10.2\n10.3")
        if st.button("🔍 Periksa Duplikat", type="primary"):
            dup = {no: n for no, n in Counter(
                   [x.strip() for x in teks.splitlines() if x.strip()]).items() if n > 1}
            if dup:
                st.error(f"Ditemukan {len(dup)} nomor ganda!")
                st.table(pd.DataFrame([(no, f"{n} kali") for no, n in dup.items()],
                                      columns=["Packing No", "Muncul"]))
            else:
                st.success("Tidak ada duplikasi. Data bersih.")

# ================= POP-UP HAPUS (MODAL) =================
_kf = st.session_state.get("konfirmasi_hapus")
if _kf:
    if _kf["tipe"] == "ship":
        konfirmasi_hapus_ship(_kf["ship"], _kf["n"])
    else:
        konfirmasi_hapus_pack(_kf["row"])

# ========================= AKHIR FILE =========================